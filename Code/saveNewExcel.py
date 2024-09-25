import pandas as pd
from openpyxl import load_workbook
import os
from util import *

def saveNewExcel(self, dataframes, file_name, save_flag) :
    file_path_exist = os.path.exists(os.getcwd()+'\\'+file_name)
    try:
        if  (file_path_exist == True and save_flag == 0):
            # Mở file Excel
            book = load_workbook(file_name)
            
            # Lấy danh sách các sheet có trong workbook
            sheet_names = book.sheetnames
            
            # Lặp qua từng sheet để xử lý dữ liệu
            with pd.ExcelWriter(file_name, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
                for sheet_name in sheet_names:
                    # Đọc sheet hiện tại
                    existing_df = pd.read_excel(file_name, sheet_name=sheet_name)
                    
                    if sheet_name in dataframes:
                        # Gộp dữ liệu mới vào dữ liệu hiện tại
                        updated_df = pd.concat([existing_df, dataframes[sheet_name]], ignore_index=True)
                        
                        # Ghi lại toàn bộ dữ liệu (cũ + mới) vào sheet
                        updated_df.to_excel(writer, sheet_name=sheet_name, index=False)
                        print(f"Đã thêm dữ liệu mới vào sheet '{sheet_name}'.")
                        update_message_status_box(self, f"Đã thêm dữ liệu mới vào sheet '{sheet_name}'.")
        else:
            # Tạo file mới và ghi dữ liệu
            with pd.ExcelWriter(file_name, engine='openpyxl') as writer:
                for sheet_name in dataframes:
                    dataframes[sheet_name].to_excel(writer, sheet_name=sheet_name, index=False)
                print(f"Đã tạo file mới '{file_name}' và ghi dữ liệu vào sheet.")
                update_message_status_box(self, f"Đã tạo file mới '{file_name}' và ghi dữ liệu vào sheet.")
    except PermissionError:
        update_message_status_box(self, f"Không thể ghi vào file {file_name}. File có thể đang mở hoặc không có quyền ghi.")
        print(f"Không thể ghi vào file {file_name}. File có thể đang mở hoặc không có quyền ghi.")
    