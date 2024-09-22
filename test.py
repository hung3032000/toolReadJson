import tkinter as tk
from tkinter import filedialog, messagebox
import pandas as pd
import threading
import pandas as pd
from openpyxl import load_workbook
import os
def process_json_file(file_path):
    try:
        # Đọc dữ liệu từ file JSON vào DataFrame
        data = pd.read_json(file_path)

        # Xử lý dữ liệu theo nhu cầu (ví dụ: hiển thị dữ liệu đầu tiên)
        name = 'SearchData2'
        excel_file_path = name + '.xlsx'
        data.to_excel(excel_file_path, index=False, engine='openpyxl')

        print(f"Excel file has been created: {excel_file_path}")



        file_path = 'SearchData2.xlsx'
        data = pd.read_excel(file_path)

        # Group case: Find INV_NO with multiple ZZ_IF_ID and BL_SRC_NO
        grouped_data = data.groupby('INV_NO').agg({
            'ZZ_IF_ID': pd.Series.nunique, 
            'BL_SRC_NO': pd.Series.nunique
        }).reset_index()

        # Filter for INV_NO where both ZZ_IF_ID and BL_SRC_NO have more than one unique value
        group_case = grouped_data[(grouped_data['ZZ_IF_ID'] > 1) & (grouped_data['BL_SRC_NO'] > 1)]

        # Show the results
        print("group case")
        print(group_case)

            
        # Split case: Find ZZ_IF_ID with multiple INV_NO
        split_data = data.groupby('ZZ_IF_ID').agg({
            'INV_NO': pd.Series.nunique, 
            'INV_CUST_CD': pd.Series.nunique
        }).reset_index()

        # Filter for ZZ_IF_ID where both INV_NO and INV_CUST_CD have more than one unique value
        split_case_manual = split_data[(split_data['INV_NO'] > 1)]

        # Show the results
        print("split case manual")
        print(split_case_manual)


        # Split case: Find ZZ_IF_ID with multiple INV_NO and INV_CUST_CD
        # Filter for ZZ_IF_ID where both INV_NO and INV_CUST_CD have more than one unique value
        split_case = split_data[(split_data['INV_NO'] > 1) & (split_data['INV_CUST_CD'] > 1)]

        # Show the results
        print("split case multipayer")
        print(split_case)


        supplement_data = data.groupby(['INV_NO', 'BL_SRC_NO']).agg({
            'ZZ_IF_ID': pd.Series.nunique
        }).reset_index()
        supplement_case = supplement_data[(supplement_data['ZZ_IF_ID'] > 1)]

        print("supplement case")
        print(supplement_case)

        supplement_data = data.groupby(['INV_NO', 'BL_SRC_NO', 'INV_CUST_CD']).agg({
            'ZZ_IF_ID': pd.Series.nunique
        }).reset_index()
        supplement_case_test = supplement_data[(supplement_data['ZZ_IF_ID'] > 1)]

        print("supplement_case_test")
        print(supplement_case_test)
        dataframes = {
            'split_case_manual': pd.concat([split_case_manual], ignore_index=True),
            'split_case': pd.concat([split_case], ignore_index=True),
            'group_case': pd.concat([group_case], ignore_index=True),
            'supplement_case_test': pd.concat([supplement_case_test], ignore_index=True)
        }


        # Tạo file Excel mới

        file_name = 'output.xlsx'
        if os.path.exists(file_name):
            # Mở file Excel
            book = load_workbook(file_name)
            
            # Lấy danh sách các sheet có trong workbook
            sheet_names = book.sheetnames
            
            # Lặp qua từng sheet để xử lý dữ liệu
            with pd.ExcelWriter(file_name, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
                for sheet_name in sheet_names:
                    # Đọc sheet hiện tại
                    existing_df = pd.read_excel(file_name, sheet_name=sheet_name)
                    
                    if sheet_name in dataframes:
                        # Gộp dữ liệu mới vào dữ liệu hiện tại
                        updated_df = pd.concat([existing_df, dataframes[sheet_name]], ignore_index=True)
                        
                        # Ghi lại toàn bộ dữ liệu (cũ + mới) vào sheet
                        updated_df.to_excel(writer, sheet_name=sheet_name, index=False)
                        print(f"Đã thêm dữ liệu mới vào sheet '{sheet_name}'.")
        else:
            # Tạo file mới và ghi dữ liệu
            with pd.ExcelWriter(file_name, engine='openpyxl') as writer:
                split_case_manual.to_excel(writer, sheet_name='split_case_manual', index=False)
                split_case.to_excel(writer, sheet_name='split_case', index=False)
                group_case.to_excel(writer, sheet_name='group_case', index=False)
                supplement_case_test.to_excel(writer, sheet_name='supplement_case_test', index=False)
            print(f"Đã tạo file mới '{file_name}' và ghi dữ liệu vào sheet.")
        
        # Sử dụng phương thức .after() để cập nhật UI từ luồng chính
        # root.after(0, lambda c=content: display_data(c))
    except Exception as e:
        root.after(0, lambda: messagebox.showerror("Lỗi", f"Không thể đọc file JSON: {e}"))

def display_data(content):
    text_box.delete(1.0, tk.END)  # Xóa nội dung hiện tại trong textbox
    text_box.insert(tk.END, content)  # Chèn nội dung mới vào textbox

def on_button_click():
    file_path = filedialog.askopenfilename(
        title="Chọn một file JSON",
        filetypes=(("JSON Files", "*.json"), ("All Files", "*.*"))
    )
    
    if file_path:
        # Tạo một luồng mới để xử lý file JSON
        threading.Thread(target=process_json_file, args=(file_path,), daemon=True).start()
    else:
        messagebox.showwarning("Cảnh báo", "Không có file nào được chọn")

# Tạo cửa sổ chính
root = tk.Tk()
root.title("Ứng dụng Tkinter")

# Tạo và đặt textbox
text_box = tk.Text(root, wrap='word', width=80, height=20)
text_box.pack(pady=20)  # Thêm khoảng cách từ trên và dưới

# Tạo và đặt nút
button = tk.Button(root, text="Chọn file JSON", command=on_button_click)
button.pack(pady=10)  # Thêm khoảng cách từ trên và dưới

# Bắt đầu vòng lặp chính của ứng dụng
root.mainloop()
